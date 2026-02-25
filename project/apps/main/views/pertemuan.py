import openpyxl
from apps.main.forms.pertemuan import PertemuanExcelForm, PertemuanForm
from apps.main.forms.presensi import PresensiExcelForm, PresensiTotalExcelForm
from apps.main.models import Pertemuan, Presensi, TipePertemuan
from apps.main.views.base import AdminRequiredMixin, CustomTemplateBaseMixin
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import AccessMixin, LoginRequiredMixin
from django.db import transaction
from django.db.models import (CharField, Count, Exists, F, IntegerField,
                              OuterRef, Q, Subquery, Value)
from django.db.models.functions import Concat, ExtractYear
from django.http.response import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.views import View
from django.views.generic import (CreateView, DeleteView, FormView, ListView,
                                  TemplateView, UpdateView)


# =====================================================================================================
#                                              BASE LOAD PAGE
# =====================================================================================================
class BasePertemuanListView(CustomTemplateBaseMixin, TemplateView):

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        context['data_tipe_pertemuan'] = TipePertemuan.objects.annotate(jumlah_pertemuan=Count('pertemuan', filter=Q(pertemuan__akhir__gte=now)))
        # context['data_tahun_pertemuan'] = Pertemuan.objects.values_list('mulai__year', flat=True).distinct().order_by('-mulai__year')
        print(Pertemuan.objects.all())
        # context['data_tahun_pertemuan'] = (
        #     Pertemuan.objects
        #     .annotate(tahun=ExtractYear('mulai'))
        #     .values_list('tahun', flat=True)
        #     .exclude(tahun__isnull=True)
        #     .distinct().order_by('-tahun')
        # )
        context['data_tahun_pertemuan'] = (
            Pertemuan.objects
            .values_list('mulai__year', flat=True)
            .exclude(mulai__isnull=True)
            .distinct()
            .order_by('-mulai__year')
        )
        print(context['data_tahun_pertemuan'])
        context['tipe_pertemuan'] = int(self.request.GET.get('tipe_pertemuan', 1))
        context['tahun_pertemuan'] = context['data_tahun_pertemuan'][0] if context['data_tahun_pertemuan'] else ''
        return context

# =====================================================================================================
#                                              ADMIN LOAD PAGE
# =====================================================================================================
class AdminPertemuanListView(AdminRequiredMixin, BasePertemuanListView):
    template_name = 'main/admin/pertemuan/table.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        presensi_excel_form = PresensiExcelForm()
        presensi_excel_form.fields['tipe_pertemuan'].queryset = TipePertemuan.objects.all()
        context['presensi_excel_form'] = presensi_excel_form

        pertemuan_excel_form = PertemuanExcelForm()
        context['pertemuan_excel_form'] = pertemuan_excel_form

        presensi_total_excel_form = PresensiTotalExcelForm()
        context['presensi_total_excel_form'] = presensi_total_excel_form

        return context


class AdminPertemuanCreateView(AdminRequiredMixin, CustomTemplateBaseMixin, CreateView):
    model = Pertemuan
    template_name = 'main/admin/pertemuan/add.html'
    form_class = PertemuanForm

    def get_success_url(self):
        tipe_id = self.object.tipe_pertemuan.id
        return reverse_lazy('main:admin.pertemuan.table') + f'?tipe_pertemuan={tipe_id}'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Data AIK berhasil dibuat!")
        return response


class AdminPertemuanUpdateView(AdminRequiredMixin, CustomTemplateBaseMixin, UpdateView):
    model = Pertemuan
    template_name = 'main/admin/pertemuan/add.html'
    form_class = PertemuanForm

    def get_success_url(self):
        tipe_id = self.object.tipe_pertemuan.id
        return reverse_lazy('main:admin.pertemuan.table') + f'?tipe_pertemuan={tipe_id}'

    def get_object(self, queryset=None):
        return get_object_or_404(self.model, id=self.kwargs.get('id'))

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Data AIK berhasil diperbarui")
        return response


class AdminPertemuanExcelImportView(AdminRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        form = PertemuanExcelForm(request.POST, request.FILES)

        if not form.is_valid():
            messages.error(request, f"Form tidak valid. {form.get_form_errors()}")
            return redirect('main:admin.pertemuan.table')

        try:
            workbook = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True)
        except Exception as e:
            messages.error(request, f"Gagal membaca file Excel: {e}")
            return redirect('main:admin.pertemuan.table')

        def col_is_valid(val):
            return val not in (None, '', 0)

        try:
            with transaction.atomic():
                for sheet in workbook.worksheets:

                    if sheet.max_row < 2:
                        continue

                    for idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=3):
                        if not any(row):
                            continue

                        try:
                            ql_tanggal = row[1]
                            ql_ke = row[2]
                            tafsir_tanggal = row[3]
                            tafsir_ke = row[4]
                            tarjih_tanggal = row[5]
                            tarjih_ke = row[6]
                            webinar_tanggal = row[7]
                            webinar_ke = row[8]

                            print(f"{ql_tanggal}|{ql_ke}|{tafsir_tanggal}|{tafsir_ke}|{tarjih_tanggal}|{tarjih_ke}|{webinar_tanggal}|{webinar_ke}")

                            mapping = [
                                ('Kajian Qiyamul Lail', ql_tanggal, ql_ke),
                                ('Kajian Tafsir', tafsir_tanggal, tafsir_ke),
                                ('Kajian Tarjih', tarjih_tanggal, tarjih_ke),
                                ('Webinar', webinar_tanggal, webinar_ke),
                            ]

                            for nama, tanggal, ke in mapping:
                                if not (col_is_valid(tanggal) and col_is_valid(ke)):
                                    continue

                                tipe_pertemuan = get_object_or_404(
                                    TipePertemuan,
                                    nama__iexact=nama
                                )

                                if timezone.is_naive(tanggal):
                                    tanggal = timezone.make_aware(tanggal)

                                Pertemuan.objects.update_or_create(
                                    tipe_pertemuan=tipe_pertemuan,
                                    judul=f'Ke-{int(ke)}',
                                    defaults={
                                        'mulai': tanggal,
                                        'akhir': tanggal + timezone.timedelta(hours=2),
                                        'presensi_mulai': tanggal,
                                        'presensi_akhir': tanggal + timezone.timedelta(hours=2),
                                    }
                                )

                        except Exception as e:
                            raise Exception(f"Sheet='{sheet.title}' " f"Baris={idx} " f"Error={e}")

        except Exception as e:
            messages.error(request, f"Gagal impor Excel. {e}")
            return redirect('main:admin.pertemuan.table')

        messages.success(request, "Import selesai. Semua data kajian berhasil diunggah.")
        return redirect('main:admin.pertemuan.table')


# =====================================================================================================
#                                                ADMIN SERVICE
# =====================================================================================================
class AdminPertemuanDeleteListView(AdminRequiredMixin, View):
    def post(self, request):
        list_id = request.POST.getlist('list_id')

        if not list_id:
            messages.error(request, 'Invalid request!')
            return redirect('main:admin.pertemuan.table')

        try:
            with transaction.atomic():
                pertemuandata = Pertemuan.objects.filter(id__in=list_id)

                if not pertemuandata.exists():
                    messages.warning(request, 'Tidak ada data dipilih untuk dihapus')
                    return redirect('main:admin.pertemuan.table')

                # simpan tipe sebelum delete
                pertemuan_obj = pertemuandata.first()
                tipe_id = pertemuan_obj.tipe_pertemuan.id

                deleted_count = pertemuandata.count()
                pertemuandata.delete()

            messages.success(request, f'{deleted_count} data berhasil dihapus')

        except Exception as e:
            messages.error(request, f'Terdapat kesalahan ketika menghapus data: {str(e)}')
            return redirect('main:admin.pertemuan.table')

        return redirect(reverse_lazy('main:admin.pertemuan.table') + f'?tipe_pertemuan={tipe_id}')


# =====================================================================================================
#                                              KARYAWAN LOAD PAGE
# =====================================================================================================
class UserPertemuanListView(BasePertemuanListView):
    template_name = 'main/user/pertemuan/table.html'


# =====================================================================================================
#                                                USER SERVICE
# =====================================================================================================
class UserPertemuanListJsonView(LoginRequiredMixin, View):
    def post(self, request):
        # ------------------------------------------------------------------
        # Queryset dengan filter
        # ------------------------------------------------------------------
        queryset = Pertemuan.objects.all().order_by('-id')
        # filter tipe pertemuan
        if tipe_pertemuan_id := request.POST.get('tipe_pertemuan'):
            queryset = queryset.filter(tipe_pertemuan__id=tipe_pertemuan_id)
        # filter tahun
        if tahun_pertemuan := request.POST.get('tahun_pertemuan'):
            queryset = queryset.filter(mulai__year=tahun_pertemuan)

        # 🔑 SUBQUERY FOR PRESENSI ID
        presensi_subquery = Presensi.objects.filter(
            pertemuan=OuterRef('pk'),
            peserta=request.user
        ).values('id')[:1]

        queryset = queryset.annotate(
            ada_presensi=Count(
                'presensi',
                filter=Q(presensi__peserta=request.user)
            ),
            presensi_id=Subquery(presensi_subquery, output_field=IntegerField()),
            materi_url=Concat(
                Value(settings.MEDIA_URL),
                F('materi'),
                output_field=CharField()
            )
        )
        # ------------------------------------------------------------------
        # Output JSON
        # ------------------------------------------------------------------
        data = queryset.values(
            'id',
            'tipe_pertemuan__id',
            'tipe_pertemuan__nama',
            'tipe_pertemuan__has_sertifikat',
            'judul',
            'deskripsi',
            'pembicara',
            'mulai',
            'akhir',
            'presensi_mulai',
            'presensi_akhir',
            'tautan',
            'materi',
            'materi_url',
            'ada_presensi',
            'presensi_id',
            'sertifikat',
            'created_at',
            'updated_at'
        )

        return JsonResponse(list(data), safe=False)
