import datetime
from collections import defaultdict

import openpyxl
from apps.main.forms.presensi import (PresensiExcelForm, PresensiForm,
                                      PresensiTotalExcelForm)
from apps.main.models import Lembaga, Pertemuan, Presensi, TipePertemuan
from apps.main.views.base import (AdminRequiredMixin, CustomTemplateBaseMixin,
                                  in_grup)
from apps.services.cetak_pdf import render_to_pdf
from apps.services.stream_pdf import stream_sertifikat_pdf
from apps.services.utils import profilesync
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import AccessMixin, LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Value
from django.db.models.functions import Replace
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.http.response import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.views import View
from django.views.generic import (CreateView, DeleteView, FormView, ListView,
                                  TemplateView, UpdateView)


# =====================================================================================================
#                                              MIXINS
# =====================================================================================================
class AdminorUserPresensiRequiredMixin(AccessMixin):
    def dispatch(self, request, *args, **kwargs):
        presensi_obj = get_object_or_404(Presensi, id=kwargs.get("presensi_id"))
        if in_grup(request.user, 'admin') or presensi_obj.peserta == request.user:
            return super().dispatch(request, *args, **kwargs)
        return HttpResponseForbidden("Anda tidak memiliki hak akses.")


# =====================================================================================================
#                                              ADMIN LOAD PAGE
# =====================================================================================================
class AdminPresensiListView(AdminRequiredMixin, CustomTemplateBaseMixin, ListView):
    model = Presensi
    template_name = "main/admin/presensi/table.html"
    context_object_name = 'data_presensi'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pertemuan_id = self.kwargs.get("pertemuan_id")
        context["pertemuan"] = get_object_or_404(Pertemuan, id=pertemuan_id) if pertemuan_id else None
        return context

    def get_queryset(self):
        pertemuan_id = self.kwargs.get("pertemuan_id")
        return self.model.objects.filter(pertemuan__id=pertemuan_id).order_by('-id')


class AdminPresensiExcelImportV2View(AdminRequiredMixin, View):

    def get_or_sync_user(self, username, nip):
        """
        Cari user:
        1. username
        2. profile.nip
        3. API sync (profilesync)
        """
        user = User.objects.select_related("profile").filter(Q(username=username) | Q(profile__nip=nip)).first()

        if user:
            return user

        user, is_success = profilesync(username)

        if not is_success or not user:
            raise Exception(f"Peserta tidak ditemukan " f"(username='{username}', NIP='{nip}')")

        return user


    def _import_regular(self, tipe_pertemuan: TipePertemuan, workbook):
        user_cache = {}
        errors = []  # ⬅️ kumpulkan error

        for sheet in workbook.worksheets:

            if sheet.max_row < 2:
                continue

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                try:
                    timestamp = row[0]
                    email = str(row[1]).strip()
                    nip = str(row[4]).strip()
                    kajian = str(row[5]).strip()
                    kesimpulan = row[6]

                    if not timestamp or not email:
                        raise ValueError("Timestamp atau email kosong")

                    username = email.replace('@ums.ac.id', '').lower().strip()
                    cache_key = f"{username}:{nip}"

                    if cache_key in user_cache:
                        peserta = user_cache[cache_key]
                    else:
                        peserta = self.get_or_sync_user(username, nip)
                        user_cache[cache_key] = peserta

                    if settings.USE_TZ:
                        dt = timezone.make_aware(timestamp) if timezone.is_naive(timestamp) else timestamp
                    else:
                        dt = timestamp.replace(tzinfo=None) if timezone.is_aware(timestamp) else timestamp

                    mulai = dt.replace(minute=0, second=0, microsecond=0)

                    pertemuan_obj, _ = Pertemuan.objects.get_or_create(
                        tipe_pertemuan=tipe_pertemuan,
                        judul=kajian,
                        defaults={
                            'mulai': mulai,
                            'akhir': mulai + timezone.timedelta(hours=2),
                            'presensi_mulai': mulai,
                            'presensi_akhir': mulai + timezone.timedelta(hours=2),
                        }
                    )

                    Presensi.objects.update_or_create(
                        pertemuan=pertemuan_obj,
                        peserta=peserta,
                        defaults={
                            'rangkuman': kesimpulan,
                            'created_at': dt,
                            'updated_at': dt,
                        }
                    )

                except Exception as e:
                    errors.append(
                        f"Sheet '{sheet.title}' baris {idx}: {e}"
                    )

        return errors


    def post(self, request, *args, **kwargs):
        form = PresensiExcelForm(request.POST, request.FILES)
        form.fields['tipe_pertemuan'].queryset = TipePertemuan.objects.all()

        if not form.is_valid():
            messages.error(request, f"Form tidak valid. {form.get_form_errors()}")
            return redirect('main:admin.pertemuan.table')

        try:
            workbook = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True)
        except Exception as e:
            messages.error(request, f"Gagal membaca file Excel: {e}")
            return redirect('main:admin.pertemuan.table')

        try:
            with transaction.atomic():
                tipe = form.cleaned_data['tipe_pertemuan']
                errors = self._import_regular(tipe, workbook)

                if errors:
                    # gabungkan error jadi satu pesan
                    error_text = "<br>".join(errors)
                    messages.warning(request, mark_safe(f"Import selesai dengan error:<br>{error_text}"))
                    return redirect('main:admin.pertemuan.table')

        except Exception as e:
            messages.error(request, f"Gagal impor Excel. {e}")
            return redirect('main:admin.pertemuan.table')

        messages.success(request, "Import selesai. Semua data berhasil diunggah.")
        return redirect('main:admin.pertemuan.table')


class AdminPresensiTotalExcelImportView(AdminRequiredMixin, View):

    def _sync_presensi(self, user, pertemuan_qs, total_excel):
        pertemuan_list = list(pertemuan_qs.order_by('id'))

        if not pertemuan_list:
            return

        # Batasi total sesuai jumlah pertemuan tersedia
        total_excel = min(int(total_excel or 0), len(pertemuan_list))

        existing = Presensi.objects.filter(
            peserta=user,
            pertemuan__in=pertemuan_list
        ).order_by('id')

        current_total = existing.count()

        # Jika kurang → tambahkan
        if current_total < total_excel:
            need = total_excel - current_total

            # ambil pertemuan yang belum ada presensinya
            existing_pertemuan_ids = set(existing.values_list('pertemuan_id', flat=True))

            available = [
                p for p in pertemuan_list
                if p.id not in existing_pertemuan_ids
            ]

            for p in available[:need]:
                Presensi.objects.create(
                    peserta=user,
                    pertemuan=p,
                    rangkuman="Auto import dari total Excel 2025"
                )


    def post(self, request, *args, **kwargs):
        form = PresensiTotalExcelForm(request.POST, request.FILES)

        if not form.is_valid():
            messages.error(request, f"Form tidak valid. {form.get_form_errors()}")
            return redirect('main:admin.pertemuan.table')

        try:
            workbook = openpyxl.load_workbook(
                request.FILES['excel_file'],
                data_only=True
            )
        except Exception as e:
            messages.error(request, f"Gagal membaca file Excel: {e}")
            return redirect('main:admin.pertemuan.table')

        nip_errors = []

        try:
            with transaction.atomic():
                sheet = workbook['Sheet1']

                ql_2025 = Pertemuan.objects.filter(tipe_pertemuan__nama__iexact='Kajian Qiyamul Lail', mulai__year=2025)
                webinar_2025 = Pertemuan.objects.filter(tipe_pertemuan__nama__iexact='Webinar', mulai__year=2025)
                tarjih_2025 = Pertemuan.objects.filter(tipe_pertemuan__nama__iexact='Kajian Tarjih', mulai__year=2025)

                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

                    if not any(row):
                        continue

                    nip = row[0]
                    tahun = row[15]
                    ql = row[8]
                    webinar = row[9]
                    tarjih = row[10]

                    if tahun != 2025:
                        continue

                    print(nip, ql, webinar, tarjih)

                    queryset = User.objects.annotate(nip_normalized=Replace('profile__nip', Value('.'), Value(''))).filter(nip_normalized=str(nip))

                    if queryset.count() == 0:
                        nip_errors.append(f"Baris {idx}: NIP '{nip}' tidak ditemukan")

                    elif queryset.count() > 1:
                        nip_errors.append(f"Baris {idx}: NIP '{nip}' lebih dari 1 user")

                    else:
                        user = queryset.first()
                        # simpan data total presensi di sini
                        self._sync_presensi(user, ql_2025, ql)
                        self._sync_presensi(user, webinar_2025, webinar)
                        self._sync_presensi(user, tarjih_2025, tarjih)

                # ⬅️ tampilkan error seperti versi regular
                if nip_errors:
                    error_text = "<br>".join(nip_errors)
                    messages.warning(request, mark_safe(f"Import selesai dengan error:<br>{error_text}"))
                    return redirect('main:admin.pertemuan.table')

        except Exception as e:
            messages.error(request, f"Gagal impor Excel. {e}")
            return redirect('main:admin.pertemuan.table')

        messages.success(request, "Import selesai. Semua data total presensi berhasil diunggah.")
        return redirect('main:admin.pertemuan.table')


# =====================================================================================================
#                                              USER LOAD PAGE
# =====================================================================================================
class UserPresensiCreateView(CustomTemplateBaseMixin, View):
    template_name = "main/user/presensi/add.html"

    def dispatch(self, request, *args, **kwargs):
        self.pertemuan = get_object_or_404(Pertemuan, id=kwargs.get("id"))
        now = timezone.now()

        if request.method == "GET":
            if self.pertemuan.presensi_mulai and now < self.pertemuan.presensi_mulai:
                messages.warning(request, "Presensi belum dibuka.")
                tipe_id = self.pertemuan.tipe_pertemuan.id
                return redirect(reverse_lazy("main:user.pertemuan.table")+ f"?tipe_pertemuan={tipe_id}")

        if request.method == "POST":
            if self.pertemuan.presensi_mulai and now < self.pertemuan.presensi_mulai:
                messages.error(request, "Presensi belum dibuka.")
                return redirect(request.path)

            if self.pertemuan.presensi_akhir and now > self.pertemuan.presensi_akhir:
                messages.error(request, "Presensi sudah ditutup.")
                return redirect(request.path)

        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        presensi = Presensi.objects.filter(pertemuan=self.pertemuan, peserta=request.user).first()

        form = PresensiForm(instance=presensi)

        return render(request, self.template_name, {
            "form": form,
            "pertemuan": self.pertemuan,
            "now": timezone.now()
        })

    def post(self, request, *args, **kwargs):
        presensi = Presensi.objects.filter(pertemuan=self.pertemuan, peserta=request.user).first()

        form = PresensiForm(request.POST, request.FILES, instance=presensi)

        if not form.is_valid():
            return render(request, self.template_name, {
                "form": form,
                "pertemuan": self.pertemuan
            })

        presensi = form.save(commit=False)
        presensi.pertemuan = self.pertemuan
        presensi.peserta = request.user
        presensi.save()

        messages.success(request, "Anda berhasil melakukan absensi kehadiran.")

        tipe_id = self.pertemuan.tipe_pertemuan.id
        return redirect(reverse_lazy("main:user.pertemuan.table") + f"?tipe_pertemuan={tipe_id}")


class UserPresensiBaganView(LoginRequiredMixin, View):
    def get(self, request):
        tahun = request.GET.get('tahun')
        karyawan_id = request.GET.get('karyawan')

        if not tahun or not karyawan_id:
            return JsonResponse({"error": "Tahun dan karyawan harus disediakan."}, status=400)

        karyawan_obj = get_object_or_404(User, id=karyawan_id)
        tanggalmulaimasuk = None
        if karyawan_obj.profile.tanggalmulaimasuk:
            tanggalmulaimasuk = datetime.datetime.strptime(
                karyawan_obj.profile.tanggalmulaimasuk,
                "%Y-%m-%d"
            ).date()

        try:
            pertemuan_filter = Q(pertemuan__mulai__year=tahun)

            if tanggalmulaimasuk:
                pertemuan_filter &= Q(
                    pertemuan__mulai__date__gte=tanggalmulaimasuk
                )

            data = (
                TipePertemuan.objects
                .annotate(
                    total_pertemuan=Count(
                        'pertemuan',
                        filter=pertemuan_filter,
                        distinct=True
                    ),
                    total_presensi=Count(
                        'pertemuan__presensi',
                        filter=pertemuan_filter & Q(
                            pertemuan__presensi__peserta_id=karyawan_id
                        ),
                        distinct=True
                    )
                )
                .values(
                    'id',
                    'nama',
                    'total_pertemuan',
                    'total_presensi'
                )
                .order_by('nama')
            )

            return JsonResponse(list(data), safe=False)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


class LembagaPresensiGrafikView(LoginRequiredMixin, View):
    def get(self, request):
        tahun = request.GET.get('tahun')
        kode_lembaga = request.GET.get('lembaga')

        if not tahun or not kode_lembaga:
            return JsonResponse({"error": "Tahun dan lembaga harus disediakan."}, status=400)

        try:
            # peserta
            userdata = User.objects.select_related('profile')
            userdata = (userdata.filter(
                    profile__home_id__isnull=False,
                    profile__status="Aktif"
                )
                .exclude(profile__kepegawaian="Dosen Tidak Tetap")
            )

            # ---------------------------------------------------
            # Query Pertemuan + aggregate presensi
            # ---------------------------------------------------
            pertemuandata = Pertemuan.objects.all().order_by('mulai')
            if kode_lembaga != 'all':
                pertemuandata = pertemuandata.filter(mulai__year=tahun, presensi__peserta__profile__home_id=kode_lembaga)
                userdata = userdata.filter(profile__home_id=kode_lembaga)

            pertemuan_qs = (
                pertemuandata
                # .annotate(
                #     total_presensi=Count('presensi', distinct=True)
                # )
                .annotate(
                    total_presensi=Count(
                        'presensi',
                        filter=Q(
                            presensi__peserta__profile__status="Aktif"
                        ) & ~Q(
                            presensi__peserta__profile__kepegawaian="Dosen Tidak Tetap"
                        ),
                        distinct=True
                    )
                )
                .select_related('tipe_pertemuan')
                .order_by('tipe_pertemuan__id', 'mulai')
            )

            total_user = userdata.distinct().count()

            # ---------------------------------------------------
            # Grouping by TipePertemuan
            # ---------------------------------------------------
            result = {}

            for p in pertemuan_qs:
                tipe = p.tipe_pertemuan
                if not tipe:
                    continue

                if tipe.id not in result:
                    result[tipe.id] = {
                        "tipe_id": tipe.id,
                        "tipe_nama": tipe.nama,
                        "has_sertifikat": tipe.has_sertifikat,
                        "pertemuan": []
                    }

                result[tipe.id]["pertemuan"].append({
                    "pertemuan_id": p.id,
                    "judul": p.judul,
                    "mulai": p.mulai.isoformat() if p.mulai else None,
                    "total_presensi": p.total_presensi,
                    "total_user": total_user,
                })

            return JsonResponse(list(result.values()), safe=False)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


class LembagaPresensiPieView(LoginRequiredMixin, View):
    def get(self, request):
        tahun = request.GET.get('tahun')

        if not tahun:
            return JsonResponse({"error": "Tahun harus disediakan."}, status=400)

        try:
            # Build effective lembaga map: home_id → display label (jenis_id 1 or 3 only)
            lembaga_map = {}
            for lb in Lembaga.objects.select_related('superunit', 'superunit__superunit').all():
                if lb.jenis_id in [1, 3]:
                    lembaga_map[lb.kode_lembaga] = lb.namasingkat or lb.nama
                else:
                    sup = lb.superunit
                    if sup and sup.jenis_id in [1, 3]:
                        lembaga_map[lb.kode_lembaga] = sup.namasingkat or sup.nama
                    elif sup and sup.superunit and sup.superunit.jenis_id in [1, 3]:
                        lembaga_map[lb.kode_lembaga] = sup.superunit.namasingkat or sup.superunit.nama
                    else:
                        lembaga_map[lb.kode_lembaga] = lb.namasingkat or lb.nama

            presensi_qs = (
                Presensi.objects
                .filter(
                    pertemuan__mulai__year=tahun,
                    peserta__profile__status="Aktif",
                    peserta__profile__home_id__isnull=False,
                )
                .exclude(peserta__profile__kepegawaian="Dosen Tidak Tetap")
                .select_related('pertemuan__tipe_pertemuan', 'peserta__profile')
            )

            tipe_data = defaultdict(lambda: defaultdict(int))
            tipe_names = {}

            for p in presensi_qs:
                tipe = p.pertemuan.tipe_pertemuan
                if not tipe:
                    continue
                home_id = p.peserta.profile.home_id
                label = lembaga_map.get(home_id, home_id)
                tipe_data[tipe.id][label] += 1
                tipe_names[tipe.id] = tipe.nama

            output = []
            for tipe_id in sorted(tipe_data.keys()):
                counts = tipe_data[tipe_id]
                total = sum(counts.values())
                slices = sorted(
                    [
                        {
                            "lembaga": k,
                            "count": v,
                            "percent": round(v / total * 100, 1) if total else 0,
                        }
                        for k, v in counts.items()
                    ],
                    key=lambda x: -x["count"],
                )
                output.append({
                    "tipe_id": tipe_id,
                    "tipe_nama": tipe_names[tipe_id],
                    "total": total,
                    "data": slices,
                })

            return JsonResponse(output, safe=False)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


class UserPresensiSertifikatView(AdminorUserPresensiRequiredMixin, CustomTemplateBaseMixin, View):
    def get(self, request, *args, **kwargs):
        presensi_obj = get_object_or_404(Presensi, id=kwargs.get("presensi_id"))

        context_data = {
            "nip1": presensi_obj.peserta.profile.nip,
            "nama1": presensi_obj.peserta.get_full_name(),
            "homebase": presensi_obj.peserta.profile.homebase,
            "nip2": presensi_obj.peserta.profile.nip,
            "nama2": presensi_obj.peserta.get_full_name(),
        }

        pdf_stream = stream_sertifikat_pdf(
            template_path=presensi_obj.pertemuan.sertifikat.path,
            position_data=presensi_obj.pertemuan.sertifikat_position,
            context_data=context_data,
        )

        response = HttpResponse(
            pdf_stream.getvalue(),
            content_type="application/pdf"
        )

        # 🔑 inline = preview, attachment = download
        response["Content-Disposition"] = 'inline; filename="sertifikat-preview.pdf"'

        return response


class UserPresensiPresentaseView(LoginRequiredMixin, View):
    def get(self, request):
        # =========================
        # PARAMS
        # =========================
        tahun = request.GET.get('tahun')
        lembaga_id = request.GET.get('lembaga')
        karyawan_id = request.GET.get('karyawan')

        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 10))

        if not tahun:
            return JsonResponse({"error": "Tahun wajib diisi"}, status=400)

        tahun = int(tahun)

        # =========================
        # BASE QUERY: PERTEMUAN
        # =========================
        pertemuan_qs = (
            Pertemuan.objects
            .select_related('tipe_pertemuan')
            .filter(mulai__year=tahun)
        )
        pertemuan_list = list(pertemuan_qs)

        # =========================
        # TIPE LIST
        # =========================
        tipe_list = list(TipePertemuan.objects.all())

        # =========================
        # USERS QUERY
        # =========================
        users_qs = User.objects.select_related('profile')
        users_qs = (users_qs.filter(
                profile__home_id__isnull=False,
                profile__status="Aktif"
            )
            .exclude(profile__kepegawaian="Dosen Tidak Tetap"))

        if lembaga_id:
            users_qs = users_qs.filter(profile__home_id=lembaga_id)

        if karyawan_id:
            users_qs = users_qs.filter(id=karyawan_id)

        # =========================
        # PAGINATION
        # =========================
        paginator = Paginator(users_qs, page_size)
        page_obj = paginator.get_page(page)

        users = page_obj.object_list

        # =========================
        # PREFETCH PRESENSI
        # =========================
        presensi_qs = (
            Presensi.objects
            .select_related('pertemuan__tipe_pertemuan')
            .filter(pertemuan__mulai__year=tahun)
        )

        users = users.prefetch_related(
            Prefetch('presensi_set', queryset=presensi_qs)
        )

        # =========================
        # BUILD DATA
        # =========================
        data = []

        for user in users:
            profile = getattr(user, 'profile', None)

            tanggalmulaimasuk = self.parse_tanggal_masuk(
                profile.tanggalmulaimasuk if profile else None
            )

            # filter pertemuan per user
            if tanggalmulaimasuk:
                pertemuan_user = [
                    p for p in pertemuan_list
                    if p.mulai and p.mulai.date() >= tanggalmulaimasuk
                ]
            else:
                pertemuan_user = pertemuan_list

            pertemuan_ids = {p.id for p in pertemuan_user}

            # total
            total_per_tipe = defaultdict(int)
            for p in pertemuan_user:
                if p.tipe_pertemuan_id:
                    total_per_tipe[p.tipe_pertemuan_id] += 1

            # diikuti
            diikuti_per_tipe = defaultdict(int)
            seen_pertemuan = set()

            for presensi in user.presensi_set.all():
                pertemuan = presensi.pertemuan

                if not pertemuan or not pertemuan.tipe_pertemuan_id:
                    continue

                if pertemuan.id not in pertemuan_ids:
                    continue

                if pertemuan.id in seen_pertemuan:
                    continue

                seen_pertemuan.add(pertemuan.id)

                diikuti_per_tipe[pertemuan.tipe_pertemuan_id] += 1

            # build row
            row = {
                'nip': getattr(profile, 'nip', user.username),
                'nama': user.get_full_name(),
                'homebase': getattr(profile, 'homebase', '') if profile else ''
            }

            for tipe in tipe_list:
                key = tipe.nama.lower().replace(' ', '_')

                total = total_per_tipe.get(tipe.id, 0)
                diikuti = diikuti_per_tipe.get(tipe.id, 0)

                persen = round((diikuti / total) * 100, 2) if total > 0 else 0
                persen = min(persen, 100)

                row[f'{key}_persen'] = persen
                row[f'{key}_total'] = total
                row[f'{key}_diikuti'] = diikuti

            data.append(row)

        # =========================
        # RESPONSE
        # =========================
        return JsonResponse({
            'data': data,
            'pagination': {
                'page': page_obj.number,
                'page_size': page_size,
                'total': paginator.count,
                'num_pages': paginator.num_pages,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            }
        })

    def parse_tanggal_masuk(self, value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            return None


class UserPresensiPresentasePDFView(LoginRequiredMixin, View):
    def get(self, request):
        tahun = request.GET.get('tahun')
        lembaga_id = request.GET.get('lembaga')
        karyawan_id = request.GET.get('karyawan')

        if not tahun:
            return JsonResponse({"error": "Tahun wajib diisi"}, status=400)

        tahun = int(tahun)

        pertemuan_qs = (
            Pertemuan.objects
            .select_related('tipe_pertemuan')
            .filter(mulai__year=tahun)
        )
        pertemuan_list = list(pertemuan_qs)

        tipe_list = list(TipePertemuan.objects.all())

        users_qs = (
            User.objects
            .select_related('profile')
            .filter(
                profile__home_id__isnull=False,
                profile__status="Aktif"
            )
            .exclude(profile__kepegawaian="Dosen Tidak Tetap")
        )

        if lembaga_id:
            users_qs = users_qs.filter(profile__home_id=lembaga_id)

        if karyawan_id:
            users_qs = users_qs.filter(id=karyawan_id)

        presensi_qs = (
            Presensi.objects
            .select_related('pertemuan__tipe_pertemuan')
            .filter(pertemuan__mulai__year=tahun)
        )

        users = users_qs.prefetch_related(
            Prefetch('presensi_set', queryset=presensi_qs)
        )

        data = []
        for user in users:
            profile = getattr(user, 'profile', None)
            tanggalmulaimasuk = self._parse_tanggal_masuk(
                profile.tanggalmulaimasuk if profile else None
            )

            if tanggalmulaimasuk:
                pertemuan_user = [
                    p for p in pertemuan_list
                    if p.mulai and p.mulai.date() >= tanggalmulaimasuk
                ]
            else:
                pertemuan_user = pertemuan_list

            pertemuan_ids = {p.id for p in pertemuan_user}

            total_per_tipe = defaultdict(int)
            for p in pertemuan_user:
                if p.tipe_pertemuan_id:
                    total_per_tipe[p.tipe_pertemuan_id] += 1

            diikuti_per_tipe = defaultdict(int)
            seen_pertemuan = set()

            for presensi in user.presensi_set.all():
                pertemuan = presensi.pertemuan
                if not pertemuan or not pertemuan.tipe_pertemuan_id:
                    continue
                if pertemuan.id not in pertemuan_ids:
                    continue
                if pertemuan.id in seen_pertemuan:
                    continue
                seen_pertemuan.add(pertemuan.id)
                diikuti_per_tipe[pertemuan.tipe_pertemuan_id] += 1

            row = {
                'nip': getattr(profile, 'nip', user.username),
                'nama': user.get_full_name(),
                'homebase': getattr(profile, 'homebase', '') if profile else '',
            }

            for tipe in tipe_list:
                key = tipe.nama.lower().replace(' ', '_')
                total = total_per_tipe.get(tipe.id, 0)
                diikuti = diikuti_per_tipe.get(tipe.id, 0)
                persen = round((diikuti / total) * 100, 2) if total > 0 else 0
                row[f'{key}_persen'] = min(persen, 100)
                row[f'{key}_total'] = total
                row[f'{key}_diikuti'] = diikuti

            data.append(row)

        # ── pre-process rows untuk template ───────────────────────
        n_tipe = len(tipe_list)
        rows = []
        for row in data:
            tipe_data = []
            for tipe in tipe_list:
                key = tipe.nama.lower().replace(' ', '_')
                tipe_data.append({
                    'persen':  row.get(f'{key}_persen', 0),
                    'total':   row.get(f'{key}_total', 0),
                    'diikuti': row.get(f'{key}_diikuti', 0),
                })
            rows.append({
                'nip':      row['nip'],
                'nama':     row['nama'],
                'homebase': row['homebase'],
                'tipe_data': tipe_data,
            })

        # ── lebar kolom dalam mm (A4 landscape: 297 - 24mm margin = 273mm) ──
        PAGE_W = 273

        t1_no, t1_nip, t1_nama, t1_homebase = 8, 22, 68, 55
        t1_tipe_w = round((PAGE_W - t1_no - t1_nip - t1_nama - t1_homebase) / n_tipe, 1) if n_tipe else 20

        t2_no, t2_nama = 8, 75
        t2_col_w = round((PAGE_W - t2_no - t2_nama) / (n_tipe * 2), 1) if n_tipe else 20

        context = {
            'tahun':      tahun,
            'tipe_list':  tipe_list,
            'rows':       rows,
            'col_count_1': 4 + n_tipe,
            'col_count_2': 2 + n_tipe * 2,
            # table 1 widths (mm)
            't1_no': t1_no, 't1_nip': t1_nip, 't1_nama': t1_nama,
            't1_homebase': t1_homebase, 't1_tipe_w': t1_tipe_w,
            # table 2 widths (mm)
            't2_no': t2_no, 't2_nama': t2_nama, 't2_col_w': t2_col_w,
        }

        # ── render HTML → PDF via xhtml2pdf ───────────────────────
        template = 'main/pdf/presensi_aik.html'
        filename = f'Presensi_AIK_{tahun}.pdf'
        # return render(request, template, context)
        return render_to_pdf(
            template,
            context,
            filename=filename
        )

    def _parse_tanggal_masuk(self, value):
        if not value:
            return None
        try:
            return datetime.datetime.strptime(value, "%Y-%m-%d").date()
        except Exception:
            return None
